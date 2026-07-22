"""
SmartLDS Flask 后端 — 物流单证智能识别 API

启动: python app.py  (默认 http://localhost:5000)
依赖: pip install flask flask-cors openpyxl pyyaml

API 端点:
  POST /api/upload          — 上传 PDF/图片
  POST /api/recognize/<id>  — 触发识别流水线
  GET  /api/result/<id>     — 获取结构化结果
  POST /api/correct/<id>    — 保存人工校正
  GET  /api/export/<id>     — 导出 Excel / JSON
  GET  /api/image/<id>      — 获取原始图片
"""

import os
import sys
import uuid
import json
import shutil
import tempfile
import zipfile
import copy
import re
from datetime import datetime
from pathlib import Path, PurePosixPath

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename

# 确保 backend 模块可导入
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from ocr_engine import OCREngine
from preprocess import Preprocessor
from layout_parser import LayoutParser
from field_extractor import FieldExtractor
from template_signature import build_anchor_layout_signature
from vision_fallback import (
    VisionFallbackClient,
    add_meta_warning,
    attach_quality_meta,
    default_vision_settings,
    evaluate_recognition_quality,
    merge_vision_fallback_result,
    normalize_vision_model,
    probe_vision_models,
    provider_defaults,
    public_vision_settings,
    vision_settings_options,
)

# ================================================================
# Flask 应用初始化
# ================================================================

app = Flask(__name__)
CORS(app)  # 允许前端跨域请求

app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024  # 最大 32MB
app.config["UPLOAD_FOLDER"] = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "uploads"
)

# ================================================================
# 全局引擎实例（单例，避免重复加载模型）
# ================================================================

_engine = None
_preprocessor = None
_parser = None
_extractor = None
_vision_fallback = None
JOB_ID_RE = re.compile(r"^[0-9a-f]{12}$")
MAX_ZIP_MEMBERS = 300
MAX_ZIP_UNCOMPRESSED = 256 * 1024 * 1024

# 任务状态存储（内存字典，重启丢失，课设够用）
_jobs: dict[str, dict] = {}


def get_engine():
    global _engine
    if _engine is None:
        _engine = OCREngine()
    return _engine


def get_preprocessor():
    global _preprocessor
    if _preprocessor is None:
        _preprocessor = Preprocessor()
    return _preprocessor


def get_parser():
    global _parser
    if _parser is None:
        _parser = LayoutParser()
    return _parser


def get_extractor():
    global _extractor
    if _extractor is None:
        _extractor = FieldExtractor()
    return _extractor


def get_vision_fallback():
    global _vision_fallback
    if _vision_fallback is None:
        _vision_fallback = VisionFallbackClient(settings=load_vision_settings(include_secret=True))
    return _vision_fallback


def reset_vision_fallback():
    global _vision_fallback
    _vision_fallback = None


def is_valid_job_id(job_id: str) -> bool:
    return bool(JOB_ID_RE.fullmatch(job_id or ""))


def _uploads_root() -> Path:
    return Path(app.config["UPLOAD_FOLDER"]).resolve()


def vision_settings_path() -> str:
    root = _uploads_root()
    root.mkdir(parents=True, exist_ok=True)
    return str(_ensure_under_root(root / "vision_settings.json", root))


def _vision_provider_keys() -> set[str]:
    return {str(item.get("key")) for item in vision_settings_options().get("providers", [])}


def _normalize_vision_provider(provider: str | None, base_url: str | None = None) -> str:
    provider = str(provider or "qwen").strip()
    base = str(base_url or "").lower()
    if provider == "custom" and "localhost:11434" in base:
        return "ollama"
    if provider not in _vision_provider_keys():
        return "qwen"
    return provider


def _normalize_vision_base_url(provider: str, base_url: str | None = None) -> str:
    provider_cfg = provider_defaults(provider)
    base = str(base_url or provider_cfg["default_base_url"]).strip().rstrip("/")
    if provider == "ollama":
        for suffix in ("/v1/chat/completions", "/chat/completions", "/v1", "/api/chat", "/api/tags"):
            if base.endswith(suffix):
                base = base[: -len(suffix)]
                break
    return base or provider_cfg["default_base_url"]


def _default_vision_profile(provider: str) -> dict:
    provider_cfg = provider_defaults(provider)
    return {
        "model": provider_cfg["default_model"],
        "base_url": provider_cfg["default_base_url"],
        "api_key": "",
        "model_api_keys": {},
    }


def _normalize_vision_profile(provider: str, profile: dict | None = None) -> dict:
    data = profile or {}
    normalized = _default_vision_profile(provider)
    normalized.update({k: v for k, v in data.items() if v is not None})
    normalized["model"] = normalize_vision_model(provider, normalized.get("model"))
    normalized["base_url"] = _normalize_vision_base_url(provider, normalized.get("base_url"))
    normalized["api_key"] = str(normalized.get("api_key") or "")
    model_api_keys = normalized.get("model_api_keys") or {}
    normalized["model_api_keys"] = {
        str(model): str(api_key)
        for model, api_key in model_api_keys.items()
        if str(api_key or "")
    } if isinstance(model_api_keys, dict) else {}
    model_key = normalized["model_api_keys"].get(normalized["model"])
    if model_key:
        normalized["api_key"] = model_key
    return normalized


def _load_vision_profiles(saved: dict) -> dict:
    profiles: dict[str, dict] = {}
    raw_profiles = saved.get("profiles")
    if isinstance(raw_profiles, dict):
        for raw_provider, raw_profile in raw_profiles.items():
            provider = _normalize_vision_provider(raw_provider, (raw_profile or {}).get("base_url") if isinstance(raw_profile, dict) else "")
            if isinstance(raw_profile, dict):
                profiles[provider] = _normalize_vision_profile(provider, raw_profile)

    if not profiles:
        provider = _normalize_vision_provider(saved.get("provider"), saved.get("base_url"))
        profiles[provider] = _normalize_vision_profile(provider, {
            "model": saved.get("model"),
            "base_url": saved.get("base_url"),
            "api_key": saved.get("api_key", ""),
        })
    return profiles


def load_vision_settings(include_secret: bool = False) -> dict:
    settings = default_vision_settings()
    saved = read_json_file(vision_settings_path(), {}) or {}
    profiles = _load_vision_profiles(saved if isinstance(saved, dict) else {})
    provider = _normalize_vision_provider(
        (saved or {}).get("provider") if isinstance(saved, dict) else None,
        (saved or {}).get("base_url") if isinstance(saved, dict) else None,
    )
    if provider not in profiles:
        profiles[provider] = _default_vision_profile(provider)
    profile = _normalize_vision_profile(provider, profiles.get(provider))

    settings.update({
        "enabled": bool((saved or {}).get("enabled", settings.get("enabled", False))) if isinstance(saved, dict) else False,
        "provider": provider,
        "model": profile["model"],
        "base_url": profile["base_url"],
        "api_key": profile["api_key"],
        "profiles": profiles,
    })

    try:
        settings["threshold"] = float(settings.get("threshold", 0.55))
        if isinstance(saved, dict) and "threshold" in saved:
            settings["threshold"] = float(saved.get("threshold", settings["threshold"]))
    except (TypeError, ValueError):
        settings["threshold"] = 0.55

    if include_secret:
        return settings
    return public_vision_settings(settings)


def save_vision_settings(data: dict) -> dict:
    current = load_vision_settings(include_secret=True)
    current_provider = str(current.get("provider") or "qwen")
    provider = _normalize_vision_provider(data.get("provider") or current_provider, data.get("base_url"))
    provider_cfg = provider_defaults(provider)
    profiles = {
        key: _normalize_vision_profile(key, value)
        for key, value in (current.get("profiles") or {}).items()
        if key in _vision_provider_keys()
    }
    profile = _normalize_vision_profile(provider, profiles.get(provider))
    model = normalize_vision_model(provider, data.get("model") or profile.get("model") or provider_cfg["default_model"])
    model_api_keys = dict(profile.get("model_api_keys") or {})
    api_key = str(model_api_keys.get(model) or profile.get("api_key") or "")

    incoming_api_key = data.get("api_key", None)
    if incoming_api_key is not None and str(incoming_api_key).strip():
        api_key = str(incoming_api_key).strip()
        model_api_keys[model] = api_key
    if provider_cfg.get("requires_api_key") is False and incoming_api_key is None:
        api_key = ""

    profile.update({
        "model": model,
        "base_url": _normalize_vision_base_url(provider, data.get("base_url") or profile.get("base_url") or provider_cfg["default_base_url"]),
        "api_key": api_key,
        "model_api_keys": model_api_keys,
    })
    profiles[provider] = profile

    updated = {
        "enabled": bool(data.get("enabled", current.get("enabled", False))),
        "provider": provider,
        "model": profile["model"],
        "base_url": profile["base_url"],
        "api_key": profile["api_key"],
        "profiles": profiles,
        "threshold": data.get("threshold", current.get("threshold", 0.55)),
    }

    try:
        updated["threshold"] = max(0.0, min(1.0, float(updated["threshold"])))
    except (TypeError, ValueError):
        updated["threshold"] = 0.55

    write_json_file(vision_settings_path(), updated)
    reset_vision_fallback()
    return public_vision_settings(updated)


def clear_vision_settings() -> dict:
    path = vision_settings_path()
    if os.path.exists(path):
        os.remove(path)
    reset_vision_fallback()
    return public_vision_settings(default_vision_settings())


def _ensure_under_root(path: Path, root: Path) -> Path:
    resolved = path.resolve()
    if os.path.commonpath([str(root), str(resolved)]) != str(root):
        raise ValueError("path escapes upload directory")
    return resolved


def job_dir(job_id: str, create: bool = True) -> str:
    """每个任务的独立目录"""
    if not is_valid_job_id(job_id):
        raise ValueError("invalid job_id")
    root = _uploads_root()
    d = _ensure_under_root(root / job_id, root)
    if create:
        d.mkdir(parents=True, exist_ok=True)
    return str(d)


def read_json_file(path: str, default=None):
    if not os.path.exists(path):
        return default
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def write_json_file(path: str, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def config_yaml_path() -> str:
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.yaml")


def _unique_key(base: str, used: set[str]) -> str:
    key = (base or "field").strip() or "field"
    if key not in used:
        used.add(key)
        return key
    idx = 2
    while f"{key}__{idx}" in used:
        idx += 1
    unique = f"{key}__{idx}"
    used.add(unique)
    return unique


def _normalize_rect_value(value) -> list[float] | None:
    if not isinstance(value, list) or len(value) != 4:
        return None
    try:
        rect = [float(item) for item in value]
    except (TypeError, ValueError):
        return None
    x1, y1, x2, y2 = rect
    if x2 <= x1 or y2 <= y1:
        return None
    return rect


def _normalize_offset_value(value) -> dict | None:
    if not isinstance(value, dict):
        return None
    normalized = {}
    for key in ("dx", "dy", "tolerance_x", "tolerance_y"):
        try:
            normalized[key] = float(value.get(key))
        except (TypeError, ValueError):
            return None
    return normalized


def normalize_corrections_payload(raw) -> dict:
    """兼容旧 flat corrections，并归一化人工字段/表格补丁。"""
    if not isinstance(raw, dict):
        return {"fields": {}, "field_labels": {}, "manual_fields": [], "excluded_fields": [], "table_patch": None}

    is_new_shape = any(k in raw for k in ("fields", "field_labels", "manual_fields", "excluded_fields", "table_patch"))
    field_corrections = dict(raw.get("fields") or {}) if is_new_shape else dict(raw)
    field_labels = {}
    if is_new_shape and isinstance(raw.get("field_labels"), dict):
        field_labels = {
            str(key).strip(): str(value).strip()
            for key, value in raw.get("field_labels", {}).items()
            if str(key).strip() and str(value).strip()
        }

    manual_fields = []
    used_manual_keys = set()
    for item in (raw.get("manual_fields") or []):
        if not isinstance(item, dict):
            continue
        label = str(item.get("label") or item.get("key") or "").strip()
        value = str(item.get("value") or "").strip()
        if not label:
            continue
        key = str(item.get("key") or label).strip()
        key = _unique_key(key, used_manual_keys)
        manual = {"key": key, "label": label, "value": value}
        anchor_text = str(item.get("anchor_text") or "").strip()
        if anchor_text:
            manual["anchor_text"] = anchor_text
        anchor_rect = _normalize_rect_value(item.get("anchor_rect"))
        value_rect = _normalize_rect_value(item.get("value_rect"))
        if anchor_rect:
            manual["anchor_rect"] = anchor_rect
        if value_rect:
            manual["value_rect"] = value_rect
        position = str(item.get("position") or "").strip().lower()
        if position in {"right", "below", "inline"}:
            manual["position"] = position
        learned_offset = _normalize_offset_value(item.get("learned_value_offset"))
        if learned_offset:
            manual["learned_value_offset"] = learned_offset
        manual_fields.append(manual)

    excluded_fields = []
    seen_excluded = set()
    for key in raw.get("excluded_fields") or []:
        text = str(key or "").strip()
        if text and text not in seen_excluded:
            excluded_fields.append(text)
            seen_excluded.add(text)

    table_patch = raw.get("table_patch") if is_new_shape else None
    normalized_table = None
    if isinstance(table_patch, dict):
        headers = [str(h).strip() for h in (table_patch.get("headers") or []) if str(h).strip()]
        rows = []
        for row in (table_patch.get("rows") or []):
            if not isinstance(row, list):
                continue
            normalized = [str(cell) for cell in row]
            if len(normalized) < len(headers):
                normalized.extend([""] * (len(headers) - len(normalized)))
            rows.append(normalized[:len(headers)] if headers else normalized)
        if headers or rows:
            normalized_table = {
                "mode": "replace",
                "headers": headers,
                "rows": rows,
            }

    return {
        "fields": field_corrections,
        "field_labels": field_labels,
        "manual_fields": manual_fields,
        "excluded_fields": excluded_fields,
        "table_patch": normalized_table,
    }


def _merge_unique_strings(existing, incoming, limit: int | None = None) -> list[str]:
    merged = []
    seen = set()
    for value in list(existing or []) + list(incoming or []):
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        merged.append(text)
        seen.add(text)
        if limit and len(merged) >= limit:
            break
    return merged


def _field_feedback_anchors(field_name: str, info: dict) -> list[str]:
    anchors = []
    for key in ("anchor", "anchor_text"):
        value = str((info or {}).get(key) or "").strip()
        if value:
            anchors.append(value)
    label = str((info or {}).get("label") or field_name).strip()
    anchors.extend([label, field_name])
    return _merge_unique_strings([], anchors, limit=5)


def _feedback_schema_field_name(field_name: str, info: dict) -> str:
    """Use the user's corrected display name as the learned schema name when safe."""
    original = str(field_name or "").strip()
    label = str((info or {}).get("label") or "").strip()
    if not label or label == original:
        return original
    if (info or {}).get("label_corrected"):
        return label
    source = str((info or {}).get("source") or "").strip()
    status = str((info or {}).get("status") or "").strip()
    anchor_text = str((info or {}).get("anchor_text") or (info or {}).get("anchor") or "").strip()
    if source == "manual" or status == "manual_added":
        if not anchor_text or label != anchor_text:
            return label
    return original


def _field_has_reliable_anchor(field_name: str, info: dict) -> bool:
    if _normalize_rect_value((info or {}).get("anchor_rect")) and _normalize_rect_value((info or {}).get("value_rect")):
        return True
    anchors = _field_feedback_anchors(field_name, info)
    label_like = {str(field_name or "").strip(), str((info or {}).get("label") or "").strip()}
    explicit = [a for a in anchors if a and a not in label_like]
    return bool(explicit) and (info or {}).get("source") != "manual"


def _feedback_layout_signature(blocks, target_template, selected_fields, final_fields, image_size):
    """Build a vocabulary-free signature for a newly created feedback template."""
    blocks = blocks or []
    if not image_size or len(image_size) < 2:
        rects = [_block_rect(block) for block in blocks]
        rects = [rect for rect in rects if rect]
        image_size = [
            max((rect[2] for rect in rects), default=1),
            max((rect[3] for rect in rects), default=1),
        ]

    observations = {}
    excluded = set()
    fields = target_template.get("fields", {})
    for info in (final_fields or {}).values():
        if not isinstance(info, dict):
            continue
        value = str(info.get("corrected") or info.get("cleaned") or info.get("value") or "").strip()
        value_block = _find_ocr_text_block(blocks, value) if value else None
        if value_block:
            excluded.add(id(value_block))
    for field_name in selected_fields:
        info = final_fields.get(field_name) or {}
        schema_name = _feedback_schema_field_name(field_name, info)
        entry = fields.get(schema_name) or fields.get(field_name) or {}
        anchors = entry.get("anchors") or _field_feedback_anchors(field_name, info)
        anchor_block = _find_anchor_text_block(blocks, anchors)
        value = str(info.get("corrected") or info.get("cleaned") or info.get("value") or "").strip()
        value_block = _find_ocr_text_block(blocks, value) if value else None
        anchor_rect = _block_rect(anchor_block)
        if not anchor_rect or anchor_block is value_block:
            continue
        observations[schema_name] = [{
            "sample_idx": 0,
            "anchor_text": str(anchor_block.get("text") or "").strip(),
            "anchor_rect": anchor_rect,
            "score": 1.0,
        }]

    return build_anchor_layout_signature(
        [blocks],
        observations,
        [image_size],
        excluded_block_ids=[excluded],
    )


def _normalize_ocr_text(text: str | None) -> str:
    return re.sub(r"\s+", "", str(text or "")).strip().lower()


def _block_rect(block: dict | None) -> list[float] | None:
    if not isinstance(block, dict):
        return None
    rect = block.get("rect")
    if isinstance(rect, list) and len(rect) == 4:
        try:
            return [float(v) for v in rect]
        except (TypeError, ValueError):
            return None
    bbox = block.get("bbox")
    if isinstance(bbox, list) and bbox:
        try:
            xs = [float(p[0]) for p in bbox if isinstance(p, list) and len(p) >= 2]
            ys = [float(p[1]) for p in bbox if isinstance(p, list) and len(p) >= 2]
            if xs and ys:
                return [min(xs), min(ys), max(xs), max(ys)]
        except (TypeError, ValueError):
            return None
    return None


def _find_ocr_text_block(blocks: list[dict] | None, text: str) -> dict | None:
    target = _normalize_ocr_text(text)
    if not target:
        return None
    candidates = []
    for block in blocks or []:
        block_text = _normalize_ocr_text(block.get("text"))
        if not block_text:
            continue
        if block_text == target:
            return block
        if target in block_text or block_text in target:
            candidates.append(block)
    return candidates[0] if candidates else None


def _find_anchor_text_block(blocks: list[dict] | None, anchors: list[str]) -> dict | None:
    normalized = [_normalize_ocr_text(anchor) for anchor in anchors if _normalize_ocr_text(anchor)]
    if not normalized:
        return None
    for block in blocks or []:
        block_text = _normalize_ocr_text(block.get("text"))
        if not block_text:
            continue
        if any(block_text == anchor or anchor in block_text for anchor in normalized):
            return block
    return None


def _infer_learned_position(anchor_rect: list[float], value_rect: list[float]) -> str:
    ax1, ay1, ax2, ay2 = anchor_rect
    bx1, by1, bx2, by2 = value_rect
    anchor_h = max(ay2 - ay1, 5.0)
    anchor_cy = (ay1 + ay2) / 2.0
    value_cy = (by1 + by2) / 2.0
    if bx1 >= ax2 - 5 and abs(value_cy - anchor_cy) <= anchor_h * 1.4:
        return "right"
    if by1 >= ay2 - 5:
        return "below"
    return "right"


def apply_ocr_feedback_learning(
    target_template: dict,
    final_fields: dict,
    selected_fields: list[str],
    blocks: list[dict] | None,
    warnings: list[str],
) -> dict:
    """把人工/校正值反查到 OCR 坐标，学习锚点到值块的稳定偏移。"""
    changes = {"applied": False, "fields": []}
    target_fields = target_template.setdefault("fields", {})

    for field_name in selected_fields:
        info = final_fields.get(field_name)
        if not isinstance(info, dict):
            continue
        schema_name = _feedback_schema_field_name(field_name, info)
        value = str(info.get("corrected") or info.get("cleaned") or info.get("value") or "").strip()
        if not value:
            continue

        if schema_name != field_name and field_name in target_fields and schema_name not in target_fields:
            target_fields[schema_name] = target_fields.pop(field_name)
        elif schema_name != field_name and field_name in target_fields and schema_name in target_fields:
            old_entry = target_fields.pop(field_name)
            if isinstance(old_entry, dict) and isinstance(target_fields.get(schema_name), dict):
                target_fields[schema_name]["anchors"] = _merge_unique_strings(
                    target_fields[schema_name].get("anchors", []),
                    old_entry.get("anchors", []),
                    limit=10,
                )

        entry = target_fields.get(schema_name)
        if not isinstance(entry, dict):
            entry = {"label": info.get("label") or schema_name, "anchors": _field_feedback_anchors(field_name, info)}
            target_fields[schema_name] = entry

        anchors = _merge_unique_strings(entry.get("anchors", []), _field_feedback_anchors(field_name, info), limit=10)
        entry["label"] = info.get("label") or schema_name
        entry["anchors"] = anchors

        anchor_rect = _normalize_rect_value(info.get("anchor_rect"))
        value_rect = _normalize_rect_value(info.get("value_rect"))
        if not anchor_rect:
            anchor_block = _find_anchor_text_block(blocks, anchors)
            anchor_rect = _block_rect(anchor_block)
        if not value_rect:
            value_block = _find_ocr_text_block(blocks, value)
            value_rect = _block_rect(value_block)
        if not anchor_rect or not value_rect:
            warnings.append(f"字段 '{field_name}' 的人工值未能反查到可靠 OCR 坐标，已保留普通字段结构")
            continue

        ax1, ay1, ax2, ay2 = anchor_rect
        bx1, by1, bx2, by2 = value_rect
        anchor_cx = (ax1 + ax2) / 2.0
        anchor_cy = (ay1 + ay2) / 2.0
        value_cx = (bx1 + bx2) / 2.0
        value_cy = (by1 + by2) / 2.0
        value_w = max(bx2 - bx1, 5.0)
        value_h = max(by2 - by1, 5.0)
        anchor_h = max(ay2 - ay1, 5.0)

        entry["position"] = _infer_learned_position(anchor_rect, value_rect)
        entry["learned_value_offset"] = {
            "dx": round(value_cx - anchor_cx, 2),
            "dy": round(value_cy - anchor_cy, 2),
            "tolerance_x": round(max(value_w * 1.8, 80.0), 2),
            "tolerance_y": round(max(value_h * 1.8, anchor_h * 1.8, 45.0), 2),
        }
        entry["learned_sample_value"] = value

        value_pattern = entry.get("value_pattern")
        if value_pattern and not re.search(str(value_pattern), value, re.IGNORECASE):
            entry.pop("value_pattern", None)
            entry.pop("validator", None)
            warnings.append(f"字段 '{field_name}' 的旧 value_pattern 与人工值不匹配，已移除以免拦截后续识别")

        if schema_name not in changes["fields"]:
            changes["fields"].append(schema_name)
        changes["applied"] = True

    return changes


def _valid_ai_position(position: str | None) -> str | None:
    position = str(position or "").strip().lower()
    return position if position in {"right", "below", "inline"} else None


def _empty_ai_changes() -> dict:
    return {"applied": False, "keywords": [], "fields": [], "table_headers": []}


def _ai_array(value, name: str, warnings: list[str]) -> list:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    warnings.append(f"AI 增强返回的 {name} 不是数组，已跳过")
    return []


def apply_ai_template_enhancement(
    target_template: dict,
    enhancement: dict,
    selected_fields: list[str],
    include_table: bool,
    warnings: list[str],
    allow_keywords: bool = True,
) -> dict:
    """把 AI 返回的版式增强建议安全合并到目标模板。"""
    selected = {str(name).strip() for name in selected_fields if str(name).strip()}
    changes = _empty_ai_changes()
    if not isinstance(enhancement, dict):
        warnings.append("AI 增强未返回有效配置对象")
        return changes

    keywords = [
        str(item).strip()
        for item in _ai_array(enhancement.get("template_keywords"), "template_keywords", warnings)
        if str(item).strip()
    ]
    if keywords and allow_keywords:
        before = list(target_template.get("keywords", []) or [])
        target_template["keywords"] = _merge_unique_strings(before, keywords, limit=12)
        changes["keywords"] = [item for item in target_template["keywords"] if item not in before]

    target_fields = target_template.setdefault("fields", {})
    for item in _ai_array(enhancement.get("fields"), "fields", warnings):
        if not isinstance(item, dict):
            continue
        field_name = str(item.get("field") or item.get("label") or "").strip()
        if not field_name:
            continue
        if selected and field_name not in selected:
            warnings.append(f"AI 增强跳过未选字段 '{field_name}'")
            continue

        entry = target_fields.setdefault(field_name, {"label": item.get("label") or field_name})
        if not isinstance(entry, dict):
            entry = {"label": item.get("label") or field_name}
            target_fields[field_name] = entry
        entry["label"] = entry.get("label") or item.get("label") or field_name

        anchors = [
            str(anchor).strip()
            for anchor in (item.get("anchors") or [])
            if str(anchor).strip()
        ]
        if anchors:
            entry["anchors"] = _merge_unique_strings(entry.get("anchors", []), anchors, limit=10)

        position = _valid_ai_position(item.get("position"))
        if position:
            entry["position"] = position

        value_pattern = str(item.get("value_pattern") or "").strip()
        if value_pattern:
            entry["value_pattern"] = value_pattern

        for flag_name in ("multi_line", "allow_shared"):
            if flag_name in item:
                entry[flag_name] = bool(item.get(flag_name))

        confidence = item.get("confidence")
        if isinstance(confidence, (int, float)):
            entry["ai_enhance_confidence"] = max(0.0, min(1.0, float(confidence)))

        if field_name not in changes["fields"]:
            changes["fields"].append(field_name)

    if include_table:
        headers = [
            str(header).strip()
            for header in _ai_array(enhancement.get("table_headers"), "table_headers", warnings)
            if str(header).strip()
        ]
        if headers:
            before = list(target_template.get("table_headers", []) or [])
            target_template["has_table"] = True
            target_template["table_headers"] = _merge_unique_strings(before, headers)
            changes["table_headers"] = [
                item for item in target_template["table_headers"] if item not in before
            ]

    for warning in _ai_array(enhancement.get("warnings"), "warnings", warnings):
        text = str(warning).strip()
        if text:
            warnings.append(f"AI 增强：{text}")

    changes["applied"] = bool(changes["keywords"] or changes["fields"] or changes["table_headers"])
    return changes


def ai_enhance_feedback_template(
    *,
    job_id: str,
    template_name: str,
    target_template: dict,
    final_result: dict,
    selected_fields: list[str],
    include_table: bool,
    warnings: list[str],
) -> dict:
    """调用视觉大模型增强反哺配置；失败时只写 warning，不中断普通反哺。"""
    client = get_vision_fallback()
    unavailable = client.unavailable_reason()
    if unavailable:
        warnings.append(f"AI 增强未执行：{unavailable}")
        return {"applied": False, "keywords": [], "fields": [], "table_headers": []}

    try:
        file_path, file_type = find_original_file(job_id)
        with tempfile.TemporaryDirectory(prefix="smartlds_ai_feedback_") as tmp_dir:
            image_path = render_first_page_for_vision(file_path, file_type, tmp_dir)
            result = client.enhance_template_config(
                image_path=image_path,
                blocks=load_blocks(job_id),
                final_result=final_result,
                template_name=template_name,
                target_template=target_template,
                selected_fields=selected_fields,
                include_table=include_table,
            )
    except Exception as exc:
        warnings.append(f"AI 增强失败，已保留普通反哺结果：{exc}")
        return {"applied": False, "keywords": [], "fields": [], "table_headers": []}

    if not result.get("success"):
        warnings.append(f"AI 增强未执行：{result.get('warning') or '模型未返回有效建议'}")
        return {"applied": False, "keywords": [], "fields": [], "table_headers": []}

    return apply_ai_template_enhancement(
        target_template,
        result.get("result", {}),
        selected_fields=selected_fields,
        include_table=include_table,
        warnings=warnings,
        allow_keywords=(target_template.get("detection") or {}).get("mode") != "anchor_layout",
    )


def _fewshot_final_result_from_sample(sample_gt: dict, learned_result: dict) -> dict:
    fields = {}
    for field_name, cfg in (learned_result.get("fields") or {}).items():
        if not isinstance(cfg, dict):
            continue
        canonical_key = cfg.get("canonical_key") or field_name
        value = sample_gt.get(canonical_key, sample_gt.get(field_name, ""))
        if isinstance(value, (list, dict)):
            value = ""
        value = str(value or "").strip()
        fields[field_name] = {
            "label": cfg.get("label") or field_name,
            "value": value,
            "cleaned": value,
            "confidence": 1.0 if value else 0.0,
            "status": "extracted" if value else "not_found",
            "anchors": cfg.get("anchors", []),
            "canonical_key": canonical_key,
        }
    return {
        "template": learned_result.get("template_name") or "fewshot_learned",
        "fields": fields,
        "table": {},
        "meta": {},
    }


def _regenerate_fewshot_yaml(learned_result: dict) -> str:
    from fewshot import FewShotLearner

    learner = object.__new__(FewShotLearner)
    return learner._generate_yaml(
        learned_result.get("template_name") or "new_template",
        learned_result.get("keywords") or [],
        learned_result.get("fields") or {},
        has_table=bool(learned_result.get("has_table")),
        table_headers=learned_result.get("table_headers") or [],
        detection=learned_result.get("detection"),
        source=learned_result.get("source") or "fewshot",
    )


def apply_ai_fewshot_enhancement(
    learned_result: dict,
    enhancement: dict,
    warnings: list[str],
) -> dict:
    """把 AI 版式建议合并回 /api/fewshot/learn 的结果，并重生成 YAML。"""
    target_template = {
        "keywords": list(learned_result.get("keywords") or []),
        "has_table": bool(learned_result.get("has_table", False)),
        "table_headers": list(learned_result.get("table_headers") or []),
        "fields": learned_result.get("fields") or {},
        "output": list((learned_result.get("fields") or {}).keys()),
    }
    selected_fields = list(target_template["fields"].keys())
    changes = apply_ai_template_enhancement(
        target_template,
        enhancement,
        selected_fields=selected_fields,
        include_table=True,
        warnings=warnings,
        allow_keywords=False,
    )

    learned_result["keywords"] = []
    learned_result["fields"] = target_template.get("fields", {})
    learned_result["has_table"] = bool(target_template.get("has_table", False))
    learned_result["table_headers"] = target_template.get("table_headers", [])
    learned_result["yaml_text"] = _regenerate_fewshot_yaml(learned_result)
    learned_result["ai_enhanced"] = bool(changes.get("applied"))
    learned_result["ai_changes"] = changes
    return changes


def ai_enhance_fewshot_learning(samples: list[tuple[str, dict]], learned_result: dict, warnings: list[str]) -> dict:
    """Few-shot 学习后可选调用视觉模型增强配置；失败不影响本地学习结果。"""
    client = get_vision_fallback()
    unavailable = client.unavailable_reason()
    if unavailable:
        warnings.append(f"AI 增强未执行：{unavailable}")
        return {"applied": False, "keywords": [], "fields": [], "table_headers": []}

    if not samples:
        warnings.append("AI 增强未执行：缺少 Few-shot 样本")
        return {"applied": False, "keywords": [], "fields": [], "table_headers": []}

    try:
        sample_path, sample_gt = samples[0]
        file_type = os.path.splitext(sample_path)[1].lstrip(".").lower() or "pdf"
        with tempfile.TemporaryDirectory(prefix="smartlds_fewshot_ai_") as tmp_dir:
            image_path = render_first_page_for_vision(sample_path, file_type, tmp_dir)
            blocks = []
            try:
                if file_type == "pdf":
                    ocr_pages = get_engine().recognize_pdf(sample_path)
                    blocks = ocr_pages[0].get("blocks", []) if ocr_pages else []
                else:
                    from PIL import Image
                    with Image.open(sample_path) as img:
                        blocks = get_engine().recognize_image(img)
            except Exception as exc:
                warnings.append(f"AI 增强 OCR 提示生成失败，继续仅用图像增强：{exc}")

            result = client.enhance_template_config(
                image_path=image_path,
                blocks=blocks,
                final_result=_fewshot_final_result_from_sample(sample_gt, learned_result),
                template_name=learned_result.get("template_name") or "fewshot_learned",
                target_template={
                    "keywords": learned_result.get("keywords", []),
                    "has_table": bool(learned_result.get("has_table", False)),
                    "table_headers": learned_result.get("table_headers", []),
                    "fields": learned_result.get("fields", {}),
                    "output": list((learned_result.get("fields") or {}).keys()),
                },
                selected_fields=list((learned_result.get("fields") or {}).keys()),
                include_table=True,
            )
    except Exception as exc:
        warnings.append(f"AI 增强失败，已保留普通 Few-shot 学习结果：{exc}")
        return {"applied": False, "keywords": [], "fields": [], "table_headers": []}

    if not result.get("success"):
        warnings.append(f"AI 增强未执行：{result.get('warning') or '模型未返回有效建议'}")
        return {"applied": False, "keywords": [], "fields": [], "table_headers": []}

    return apply_ai_fewshot_enhancement(learned_result, result.get("result", {}), warnings)


def find_original_file(job_id: str):
    d = job_dir(job_id, create=False)
    if not os.path.isdir(d):
        return None, None
    for ext in sorted(ALLOWED_EXTENSIONS):
        candidate = os.path.join(d, f"original.{ext}")
        if os.path.exists(candidate):
            return candidate, ext
    return None, None


def apply_corrections(result: dict | None, corrections: dict | None) -> dict:
    result_copy = copy.deepcopy(result or {})
    normalized = normalize_corrections_payload(corrections or {})
    clean_corrections = normalized["fields"]
    result_copy["corrections"] = clean_corrections
    result_copy["field_labels"] = normalized["field_labels"]
    result_copy["manual_fields"] = normalized["manual_fields"]
    result_copy["excluded_fields"] = normalized["excluded_fields"]
    result_copy["table_patch"] = normalized["table_patch"]

    fields = result_copy.get("fields", {})
    if isinstance(fields, dict):
        for fname, label in normalized["field_labels"].items():
            if fname in fields and isinstance(fields[fname], dict):
                fields[fname]["label"] = label
                fields[fname]["label_corrected"] = True

        for fname, corrected_val in clean_corrections.items():
            if fname in fields and isinstance(fields[fname], dict):
                fields[fname]["corrected"] = corrected_val
                fields[fname]["status"] = "corrected"

        used = set(fields.keys())
        for item in normalized["manual_fields"]:
            key = item["key"]
            if key in used:
                key = _unique_key(key, used)
            else:
                used.add(key)
            value = item.get("value", "")
            fields[key] = {
                "label": item.get("label") or key,
                "value": value,
                "cleaned": value,
                "corrected": value,
                "confidence": 1.0,
                "status": "manual_added",
                "source": "manual",
                "anchor": "",
                "rect": [0, 0, 0, 0],
                "canonical_key": None,
            }
            if item.get("anchor_text"):
                fields[key]["anchor"] = item["anchor_text"]
                fields[key]["anchor_text"] = item["anchor_text"]
            if item.get("anchor_rect"):
                fields[key]["anchor_rect"] = item["anchor_rect"]
            if item.get("value_rect"):
                fields[key]["value_rect"] = item["value_rect"]
                fields[key]["rect"] = item["value_rect"]
            if item.get("position"):
                fields[key]["position"] = item["position"]
            if item.get("learned_value_offset"):
                fields[key]["learned_value_offset"] = item["learned_value_offset"]

        for fname in normalized["excluded_fields"]:
            if fname in fields and isinstance(fields[fname], dict):
                fields[fname]["excluded"] = True

    table_patch = normalized.get("table_patch")
    if table_patch:
        source = "manual_patch"
        if result_copy.get("table", {}).get("headers") or result_copy.get("table", {}).get("rows"):
            source = "ocr_with_manual_patch"
        result_copy["table"] = {
            "headers": table_patch.get("headers", []),
            "rows": table_patch.get("rows", []),
            "source": source,
        }
        result_copy.setdefault("meta", {})["manual_table"] = True

    result_copy.setdefault("meta", {})["manual_fields_count"] = len(normalized["manual_fields"])
    return result_copy


def _final_field_value(info: dict) -> str:
    if not isinstance(info, dict):
        return ""
    for key in ("corrected", "cleaned", "value"):
        value = info.get(key)
        if value is not None and str(value).strip():
            return str(value)
    return ""


def build_field_values(fields: dict | None) -> dict:
    """生成面向人工查看/下游导入的字段→值简洁映射。"""
    values = {}
    for field_key, info in (fields or {}).items():
        if not isinstance(info, dict):
            continue
        if info.get("excluded"):
            continue
        value = _final_field_value(info)
        if not value and info.get("status") == "not_found":
            continue
        label = str(info.get("label") or field_key).strip()
        if label:
            values[label] = value
    return values


def export_options_from_args(args) -> dict:
    """解析导出选项；默认完整明细 + 简洁键值 + 表格 + 元信息。"""
    def as_bool(name: str, default: bool) -> bool:
        value = args.get(name)
        if value is None:
            return default
        return str(value).strip().lower() in {"1", "true", "yes", "on"}

    preset = str(args.get("preset") or "combined").strip().lower()
    if preset == "values":
        defaults = {"field_values": True, "field_details": False, "table": False, "meta": False}
    elif preset == "details":
        defaults = {"field_values": False, "field_details": True, "table": False, "meta": True}
    else:
        defaults = {"field_values": True, "field_details": True, "table": True, "meta": True}

    options = {
        "field_values": as_bool("field_values", defaults["field_values"]),
        "field_details": as_bool("field_details", defaults["field_details"]),
        "table": as_bool("table", defaults["table"]),
        "meta": as_bool("meta", defaults["meta"]),
    }
    if not any(options.values()):
        options["field_values"] = True
    return options


def build_export_json_payload(result: dict, options: dict) -> dict:
    """按导出选项组装 JSON，保留详细版并可附加 field_values 简洁版。"""
    fields = result.get("fields", {}) if isinstance(result, dict) else {}
    payload = {}

    if options.get("field_details"):
        payload.update(copy.deepcopy(result))
        if isinstance(payload.get("fields"), dict):
            payload["fields"] = {
                key: value
                for key, value in payload["fields"].items()
                if not (isinstance(value, dict) and value.get("excluded"))
            }
        payload.pop("excluded_fields", None)

    if options.get("field_values"):
        payload["field_values"] = build_field_values(fields)

    if options.get("table"):
        payload["table"] = copy.deepcopy(result.get("table", {}))
        if result.get("tables"):
            payload["tables"] = copy.deepcopy(result.get("tables", []))
    elif not options.get("field_details"):
        payload.pop("table", None)
        payload.pop("tables", None)

    if options.get("meta"):
        payload["meta"] = copy.deepcopy(result.get("meta", {}))
    elif not options.get("field_details"):
        payload.pop("meta", None)

    if not options.get("field_details"):
        payload.pop("fields", None)
    return payload


def strip_debug(result: dict | None) -> dict:
    result_copy = copy.deepcopy(result or {})
    result_copy.pop("debug", None)
    return result_copy


def load_job(job_id: str):
    if not is_valid_job_id(job_id):
        return None
    if job_id in _jobs:
        return _jobs[job_id]

    d = job_dir(job_id, create=False)
    if not os.path.isdir(d):
        return None

    result = read_json_file(os.path.join(d, "result.json"), None)
    corrections = read_json_file(os.path.join(d, "corrections.json"), {}) or {}
    file_path, file_type = find_original_file(job_id)

    if result is None and file_path is None:
        return None

    filename = (result or {}).get("filename") or (os.path.basename(file_path) if file_path else "")
    status = "corrected" if corrections else ("done" if result else "uploaded")
    job = {
        "id": job_id,
        "filename": filename,
        "file_type": file_type or "",
        "file_path": file_path or "",
        "status": status,
        "created_at": (result or {}).get("recognized_at", ""),
        "result": result,
        "corrections": corrections,
    }
    _jobs[job_id] = job
    return job


def load_blocks(job_id: str):
    return read_json_file(os.path.join(job_dir(job_id, create=False), "blocks.json"), []) or []


def safe_extract_allowed_zip(zip_path: str, extract_dir: str):
    root = Path(extract_dir).resolve()
    root.mkdir(parents=True, exist_ok=True)
    extracted = []
    total_size = 0

    with zipfile.ZipFile(zip_path, "r") as zf:
        members = [info for info in zf.infolist() if not info.is_dir()]
        if len(members) > MAX_ZIP_MEMBERS:
            raise ValueError(f"ZIP contains too many files, max {MAX_ZIP_MEMBERS}")

        for info in members:
            total_size += info.file_size
            if total_size > MAX_ZIP_UNCOMPRESSED:
                raise ValueError("ZIP uncompressed size is too large")

            raw_name = info.filename.replace("\\", "/")
            member_path = PurePosixPath(raw_name)
            if member_path.is_absolute() or ".." in member_path.parts:
                raise ValueError(f"unsafe ZIP member path: {info.filename}")

            ext = member_path.name.rsplit(".", 1)[-1].lower() if "." in member_path.name else ""
            if ext not in ALLOWED_EXTENSIONS:
                continue

            dest = _ensure_under_root(root.joinpath(*member_path.parts), root)
            dest.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(info, "r") as src, open(dest, "wb") as dst:
                shutil.copyfileobj(src, dst)
            extracted.append(str(dest))

    return extracted


@app.before_request
def validate_route_job_id():
    job_id = (request.view_args or {}).get("job_id")
    if job_id is not None and not is_valid_job_id(job_id):
        return jsonify({"error": "invalid job_id"}), 400


# ================================================================
# 工具函数
# ================================================================

ALLOWED_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "tiff", "bmp"}


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def is_pdf(filename: str) -> bool:
    return filename.rsplit(".", 1)[1].lower() == "pdf" if "." in filename else False


def render_first_page_for_vision(file_path: str, file_type: str, tmp_dir: str) -> str:
    """Return an image path suitable for vision fallback; render PDF page 1 when needed."""
    if file_type == "pdf":
        import fitz
        from PIL import Image

        doc = fitz.open(file_path)
        pix = doc[0].get_pixmap(dpi=200)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        doc.close()

        image_path = os.path.join(tmp_dir, "vision_page_1.png")
        img.save(image_path, format="PNG")
        return image_path
    return file_path


def vision_fallback_threshold() -> float:
    settings = load_vision_settings(include_secret=True)
    try:
        return float(os.getenv("VISION_FALLBACK_THRESHOLD", settings.get("threshold", 0.55)))
    except (TypeError, ValueError):
        return 0.55


def run_recognition_pipeline(job: dict):
    file_path = job["file_path"]
    if not os.path.exists(file_path):
        raise FileNotFoundError("文件已丢失")

    engine = get_engine()
    preprocessor = get_preprocessor()
    parser = get_parser()
    extractor = get_extractor()
    warnings = []

    if is_pdf(job["filename"]):
        ocr_results = engine.recognize_pdf(file_path)
        if not ocr_results:
            raise ValueError("PDF 没有可识别页面")
        page = ocr_results[0]
        page_count = len(ocr_results)
        if page_count > 1:
            warnings.append(f"PDF has {page_count} pages; only page 1 is processed")
        blocks = page["blocks"]
        img_size = page["image_size"]
    else:
        from PIL import Image
        with Image.open(file_path) as img:
            pp_img = preprocessor.process_image(img, engine=engine)
            blocks = engine.recognize_image(pp_img)
            img_size = pp_img.size
        page_count = 1

    regions = parser.parse(blocks, img_size)
    extracted = extractor.extract(regions, img_size, blocks=blocks)
    meta = dict(extracted.get("meta", {}))
    meta["page_count"] = page_count
    meta["processed_page"] = 1
    if warnings:
        meta["warnings"] = warnings
    extracted["meta"] = meta

    # 质量评分只依赖本地结果，默认总是执行；真正调用视觉模型还要看环境变量。
    # 这样前端/导出都能看到本地识别置信度，同时避免高置信度样本产生额外成本。
    quality = evaluate_recognition_quality(
        extracted,
        blocks=blocks,
        threshold=vision_fallback_threshold(),
    )
    extracted = attach_quality_meta(extracted, quality, extraction_source="local_rules")

    if quality.get("should_fallback"):
        # 低置信度才尝试视觉兜底。未启用或失败时保留本地结果并写入 warning，
        # 不让外部 API 状态影响基础 OCR/规则识别流程。
        fallback = get_vision_fallback()
        meta = dict(extracted.get("meta", {}))
        meta.update({
            "vision_attempted": True,
            "vision_provider": fallback.provider,
            "vision_model": fallback.model,
            "vision_endpoint_type": getattr(fallback, "endpoint_type", ""),
        })
        extracted["meta"] = meta
        unavailable = fallback.unavailable_reason()
        if unavailable:
            fallback_result = {"success": False, "warning": unavailable}
        else:
            with tempfile.TemporaryDirectory(prefix="smartlds_vision_") as tmp_dir:
                vision_image_path = render_first_page_for_vision(
                    file_path,
                    job.get("file_type", "pdf"),
                    tmp_dir,
                )
                fallback_result = fallback.extract(
                    vision_image_path,
                    blocks=blocks,
                    local_result=extracted,
                    quality=quality,
                )

        if fallback_result.get("success"):
            extracted = merge_vision_fallback_result(
                extracted,
                fallback_result.get("result", {}),
                quality,
            )
            meta = dict(extracted.get("meta", {}))
            meta.update({
                "vision_attempted": True,
                "vision_provider": fallback.provider,
                "vision_model": fallback.model,
                "vision_endpoint_type": getattr(fallback, "endpoint_type", ""),
            })
            extracted["meta"] = meta
        else:
            warning = fallback_result.get("warning") or "视觉兜底未执行，已保留本地规则结果"
            extracted = add_meta_warning(extracted, warning)
            meta = dict(extracted.get("meta", {}))
            meta["fallback_reason"] = ", ".join(quality.get("fallback_reasons", []))
            meta.update({
                "vision_attempted": True,
                "vision_provider": fallback.provider,
                "vision_model": fallback.model,
                "vision_endpoint_type": getattr(fallback, "endpoint_type", ""),
            })
            extracted["meta"] = meta
    else:
        meta = dict(extracted.get("meta", {}))
        meta["vision_attempted"] = False
        meta["vision_skipped_reason"] = "local_quality_above_threshold"
        extracted["meta"] = meta

    result_data = {
        **extracted,
        "job_id": job["id"],
        "filename": job["filename"],
        "recognized_at": datetime.now().isoformat(),
    }

    d = job_dir(job["id"])
    write_json_file(os.path.join(d, "result.json"), result_data)
    write_json_file(os.path.join(d, "blocks.json"), blocks)

    job["status"] = "done"
    job["result"] = result_data
    job.pop("error", None)
    return result_data


# ================================================================
# API: 文件上传
# ================================================================

@app.route("/api/upload", methods=["POST"])
def api_upload():
    """上传单文件 PDF 或图片，返回 job_id"""
    if "file" not in request.files:
        return jsonify({"error": "缺少 file 字段"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "文件名为空"}), 400

    if not allowed_file(file.filename):
        return jsonify({"error": f"不支持的文件类型，允许: {ALLOWED_EXTENSIONS}"}), 400

    job_id = uuid.uuid4().hex[:12]
    d = job_dir(job_id)

    filename = secure_filename(file.filename)
    ext = filename.rsplit(".", 1)[1].lower()
    saved_path = os.path.join(d, f"original.{ext}")
    file.save(saved_path)

    _jobs[job_id] = {
        "id": job_id,
        "filename": filename,
        "file_type": ext,
        "file_path": saved_path,
        "status": "uploaded",
        "created_at": datetime.now().isoformat(),
        "result": None,
        "corrections": {},
    }

    return jsonify({
        "job_id": job_id,
        "filename": filename,
        "status": "uploaded",
    }), 201


# ================================================================
# API: 触发识别
# ================================================================

@app.route("/api/recognize/<job_id>", methods=["POST"])
def api_recognize(job_id):
    """触发完整识别流水线"""
    job = load_job(job_id)
    if job is None:
        return jsonify({"error": "任务不存在"}), 404

    if not job.get("file_path") or not os.path.exists(job["file_path"]):
        return jsonify({"error": "文件已丢失"}), 404

    try:
        job["status"] = "processing"
        result_data = run_recognition_pipeline(job)
        meta = result_data.get("meta", {})

        return jsonify({
            "job_id": job_id,
            "status": "done",
            "template": result_data["template"],
            "fields_count": meta.get("fields_extracted", 0),
            "fields_total": meta.get("fields_total", 0),
            "table_rows": meta.get("table_rows", 0),
            "warnings": meta.get("warnings", []),
        })

    except Exception as e:
        job["status"] = "error"
        job["error"] = str(e)
        return jsonify({"error": f"识别失败: {str(e)}"}), 500


# ================================================================
# API: 获取结果
# ================================================================

@app.route("/api/result/<job_id>", methods=["GET"])
def api_result(job_id):
    """获取识别结果（含 bbox 标注数据）— 优先从内存，回退到磁盘"""
    job = load_job(job_id)

    # 如果内存中没有，从磁盘读取（支持重启后访问历史）
    if job is None:
        return jsonify({"error": "任务不存在"}), 404

    if job["status"] in ("uploaded", "processing"):
        return jsonify({"job_id": job_id, "status": job["status"]}), 200

    if job["status"] == "error":
        return jsonify({"job_id": job_id, "status": "error", "error": job.get("error", "")}), 500

    result = apply_corrections(job.get("result", {}), job.get("corrections", {}))
    if request.args.get("debug") not in ("1", "true", "yes"):
        result = strip_debug(result)
    result["status"] = job["status"]
    result["corrections_payload"] = job.get("corrections", {})
    result["blocks"] = load_blocks(job_id)

    return jsonify(result)


# ================================================================
# API: 人工校正
# ================================================================

@app.route("/api/correct/<job_id>", methods=["POST"])
def api_correct(job_id):
    """保存前端发来的人工校正字段值"""
    job = load_job(job_id)
    if job is None:
        return jsonify({"error": "任务不存在"}), 404

    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "请求体不能为空"}), 400

    if job.get("result") is None:
        return jsonify({"error": "尚未识别"}), 400

    # 保存校正：兼容旧 fields 格式，同时支持人工新增字段和表格补丁
    existing = normalize_corrections_payload(job.get("corrections") or {})
    incoming = normalize_corrections_payload(data)
    corrections = {
        "fields": dict(existing["fields"]),
        "field_labels": dict(existing["field_labels"]),
        "manual_fields": incoming["manual_fields"],
        "excluded_fields": incoming["excluded_fields"],
        "table_patch": incoming["table_patch"],
    }
    corrections["fields"].update(incoming["fields"])
    corrections["field_labels"].update(incoming["field_labels"])
    job["corrections"] = corrections
    job["status"] = "corrected"

    # 写入文件
    corr_path = os.path.join(job_dir(job_id), "corrections.json")
    write_json_file(corr_path, corrections)

    return jsonify({
        "job_id": job_id,
        "status": "corrected",
        "corrected_fields": list(corrections["fields"].keys()),
        "manual_fields_count": len(corrections["manual_fields"]),
        "excluded_fields_count": len(corrections["excluded_fields"]),
        "has_table_patch": bool(corrections["table_patch"]),
    })


# ================================================================
# API: 导出
# ================================================================

@app.route("/api/export/<job_id>", methods=["GET"])
def api_export(job_id):
    """导出结果，format=json|xlsx (默认 json)"""
    job = load_job(job_id)
    if job is None:
        return jsonify({"error": "任务不存在"}), 404

    if job.get("result") is None:
        return jsonify({"error": "尚未识别"}), 400

    export_format = request.args.get("format", "json").lower()
    options = export_options_from_args(request.args)

    # 合并原始结果 + 校正
    result = strip_debug(apply_corrections(job["result"], job.get("corrections", {})))
    fields = result.get("fields", {})

    if export_format == "xlsx":
        return _export_excel(job_id, result, fields, options)
    else:
        return _export_json(job_id, result, options)


def _export_json(job_id, result, options):
    """导出为 JSON 文件"""
    payload = build_export_json_payload(result, options)
    export_path = os.path.join(job_dir(job_id), "export.json")
    with open(export_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return send_file(export_path, as_attachment=True,
                     download_name=f"{job_id}_result.json",
                     mimetype="application/json")


def _export_excel(job_id, result, fields, options):
    """导出为 Excel 文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="003882", end_color="003882", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    def style_header(row):
        for cell in row:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    if options.get("field_values"):
        ws_values = wb.create_sheet("字段键值")
        ws_values.append(["字段", "值"])
        style_header(ws_values[1])
        for label, value in build_field_values(fields).items():
            ws_values.append([label, value])
        ws_values.column_dimensions["A"].width = 24
        ws_values.column_dimensions["B"].width = 40

    if options.get("field_details"):
        ws1 = wb.create_sheet("字段明细")
        ws1.append(["字段名", "原始值", "清洗值", "最终值", "状态", "置信度", "锚点文本"])
        style_header(ws1[1])

        for fname, info in fields.items():
            if info.get("excluded"):
                continue
            conf = info.get("confidence", 0)
            display_name = info.get("label") or fname
            ws1.append([
                display_name,
                info.get("value", ""),
                info.get("cleaned", ""),
                _final_field_value(info),
                info.get("status", ""),
                f"{conf:.0%}" if isinstance(conf, (int, float)) else str(conf),
                info.get("anchor") or info.get("anchor_text", ""),
            ])

        ws1.column_dimensions["A"].width = 22
        ws1.column_dimensions["B"].width = 30
        ws1.column_dimensions["C"].width = 30
        ws1.column_dimensions["D"].width = 30
        ws1.column_dimensions["E"].width = 14
        ws1.column_dimensions["F"].width = 10
        ws1.column_dimensions["G"].width = 22

    tables = result.get("tables")
    if not isinstance(tables, list) or not tables:
        table = result.get("table", {})
        tables = [table] if table and table.get("headers") else []
    tables = [table for table in tables if isinstance(table, dict) and table.get("headers")]

    if options.get("table") and tables:
        ws2 = wb.create_sheet("表格数据")
        for table_index, table in enumerate(tables):
            if table_index:
                ws2.append([])
            title = str(table.get("title") or f"表格 {table_index + 1}").strip()
            ws2.append([title])
            title_row = ws2[ws2.max_row]
            for cell in title_row:
                cell.font = header_font
            ws2.append(table["headers"])
            style_header(ws2[ws2.max_row])

            for row in table.get("rows", []):
                ws2.append(row)

        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws2.column_dimensions[col_letter].width = 18

    if options.get("meta"):
        ws3 = wb.create_sheet("元信息")
        ws3.append(["键", "值"])
        for cell in ws3[1]:
            cell.font = header_font
        meta = {
            "任务 ID": job_id,
            "文件名": result.get("filename", ""),
            "识别版式": result.get("template", ""),
            "图片尺寸": str(result.get("image_size", "")),
            "文本块数": result.get("blocks_count", 0),
            "识别时间": result.get("recognized_at", ""),
            "包含人工字段": result.get("meta", {}).get("manual_fields_count", 0),
            "包含人工表格": "是" if result.get("meta", {}).get("manual_table") else "否",
        }
        for k, v in meta.items():
            ws3.append([k, str(v)])
        ws3.column_dimensions["A"].width = 15
        ws3.column_dimensions["B"].width = 40

    if not wb.worksheets:
        ws = wb.create_sheet("字段键值")
        ws.append(["字段", "值"])
        style_header(ws[1])

    export_path = os.path.join(job_dir(job_id), "export.xlsx")
    wb.save(export_path)
    return send_file(export_path, as_attachment=True,
                     download_name=f"{job_id}_result.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ================================================================
# API: 获取原始图片
# ================================================================

@app.route("/api/image/<job_id>", methods=["GET"])
def api_image(job_id):
    """返回原始图片（PDF 转第一页 PNG）供前端 Canvas 展示"""
    job = load_job(job_id)
    if job is None:
        return jsonify({"error": "任务不存在"}), 404

    file_path = job["file_path"]
    if not os.path.exists(file_path):
        return jsonify({"error": "文件已丢失"}), 404

    ext = job.get("file_type", "pdf")

    if ext == "pdf":
        # PDF → 第一页渲染为 PNG（200 DPI 与 OCR 引擎一致，标注框才对齐）
        import fitz
        from PIL import Image
        import io

        doc = fitz.open(file_path)
        pix = doc[0].get_pixmap(dpi=200)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        doc.close()

        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        return send_file(buf, mimetype="image/png")
    else:
        return send_file(file_path, mimetype=f"image/{ext}")


# ================================================================
# API: ZIP 批量上传 + 批量识别
# ================================================================

@app.route("/api/upload/zip", methods=["POST"])
def api_upload_zip():
    """上传 ZIP 压缩包，提取文件，返回批量 job 列表"""
    if "file" not in request.files:
        return jsonify({"error": "缺少 file 字段"}), 400

    file = request.files["file"]
    if not file.filename or not file.filename.lower().endswith(".zip"):
        return jsonify({"error": "只支持 .zip 文件"}), 400

    batch_id = uuid.uuid4().hex[:12]
    batch_dir = os.path.join(app.config["UPLOAD_FOLDER"], "batch_" + batch_id)
    os.makedirs(batch_dir, exist_ok=True)

    # 解压
    zip_path = os.path.join(batch_dir, "upload.zip")
    file.save(zip_path)
    extract_dir = os.path.join(batch_dir, "files")
    os.makedirs(extract_dir, exist_ok=True)

    try:
        extracted_files = safe_extract_allowed_zip(zip_path, extract_dir)
    except (zipfile.BadZipFile, ValueError) as e:
        shutil.rmtree(batch_dir, ignore_errors=True)
        return jsonify({"error": f"ZIP 解压失败: {str(e)}"}), 400

    # 遍历解压后的文件，为每个合法文件创建 job
    jobs = []
    for src in sorted(extracted_files):
        fname = os.path.basename(src)
        ext = fname.rsplit(".", 1)[-1].lower() if "." in fname else ""
        job_id = uuid.uuid4().hex[:12]
        d = job_dir(job_id)
        dst = os.path.join(d, f"original.{ext}")
        shutil.copy2(src, dst)

        _jobs[job_id] = {
            "id": job_id,
            "filename": fname,
            "file_type": ext,
            "file_path": dst,
            "status": "uploaded",
            "batch_id": batch_id,
            "created_at": datetime.now().isoformat(),
            "result": None,
            "corrections": {},
        }
        jobs.append({"job_id": job_id, "filename": fname})

    # 清理临时文件
    os.remove(zip_path)
    shutil.rmtree(extract_dir, ignore_errors=True)

    return jsonify({
        "batch_id": batch_id,
        "count": len(jobs),
        "jobs": jobs,
    }), 201


@app.route("/api/recognize/batch", methods=["POST"])
def api_recognize_batch():
    """批量触发识别 — 逐个处理并返回结果"""
    data = request.get_json(silent=True)
    if not data or "job_ids" not in data:
        return jsonify({"error": "请求体需包含 job_ids 数组"}), 400

    job_ids = data["job_ids"]
    results = []
    for job_id in job_ids:
        job = load_job(job_id)
        if not job:
            results.append({"job_id": job_id, "status": "error", "error": "任务不存在"})
            continue

        job["status"] = "processing"
        file_path = job.get("file_path")
        if not file_path or not os.path.exists(file_path):
            job["status"] = "error"
            results.append({"job_id": job_id, "status": "error", "error": "文件丢失"})
            continue

        try:
            result_data = run_recognition_pipeline(job)
            meta = result_data.get("meta", {})
            results.append({
                "job_id": job_id,
                "filename": job["filename"],
                "status": "done",
                "template": result_data["template"],
                "fields_extracted": meta.get("fields_extracted", 0),
                "fields_total": meta.get("fields_total", 0),
                "warnings": meta.get("warnings", []),
            })
            continue

        except Exception as e:
            job["status"] = "error"
            job["error"] = str(e)
            results.append({"job_id": job_id, "status": "error", "error": str(e)})

    return jsonify({"results": results})


@app.route("/api/results/batch/<batch_id>", methods=["GET"])
def api_batch_results(batch_id):
    """获取某批次的全部结果汇总"""
    batch_jobs = [j for j in _jobs.values() if j.get("batch_id") == batch_id]
    summary = []
    for job in batch_jobs:
        summary.append({
            "job_id": job["id"],
            "filename": job.get("filename", ""),
            "status": job.get("status", ""),
            "template": job.get("result", {}).get("template", "?") if job.get("result") else "?",
            "fields_extracted": job.get("result", {}).get("meta", {}).get("fields_extracted", 0) if job.get("result") else 0,
            "fields_total": job.get("result", {}).get("meta", {}).get("fields_total", 0) if job.get("result") else 0,
        })
    return jsonify({"batch_id": batch_id, "count": len(summary), "jobs": summary})


# ================================================================
# API: 版式配置查询
# ================================================================

@app.route("/api/config", methods=["GET"])
def api_config():
    """返回当前 config.yaml 的可视化友好格式（供前端版式管理展示）"""
    extractor = get_extractor()
    config_templates = extractor.config.get("templates", {})

    templates = []
    for tname, tcfg in config_templates.items():
        if tcfg.get("hidden", False):
            continue
        tpl_fields = tcfg.get("fields", {})
        output_list = tcfg.get("output", list(tpl_fields.keys()))

        tpl_info = {
            "name": tname,
            "has_table": tcfg.get("has_table", False),
            "keywords": tcfg.get("keywords", []),
            "source": tcfg.get("source", "builtin"),
            "detection": tcfg.get("detection"),
            "fields": [],
        }
        for fname in output_list:
            fdef = tpl_fields.get(fname, {})
            f_info = {
                "key": fname,
                "label": fdef.get("label", fname),
                "canonical_key": fdef.get("canonical_key", ""),
                "anchors": fdef.get("anchors", []),
                "position": fdef.get("position", "right"),
                "validator": fdef.get("validator") or "",
            }
            tpl_info["fields"].append(f_info)
        templates.append(tpl_info)

    return jsonify({
        "templates": templates,
        "validators": extractor.validators_cfg,
    })


# ================================================================
# API: 版式管理 — 删除版式
# ================================================================

@app.route("/api/config/<template_name>", methods=["DELETE"])
def api_config_delete(template_name):
    """删除指定版式（同步更新 config.yaml + 热重载）"""
    import yaml

    config_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "config.yaml"
    )
    if not os.path.exists(config_path):
        return jsonify({"error": "config.yaml 不存在"}), 404

    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    if template_name not in cfg.get("templates", {}):
        return jsonify({"error": f"版式 '{template_name}' 不存在"}), 404

    del cfg["templates"][template_name]

    with open(config_path, "w", encoding="utf-8") as f:
        yaml.dump(cfg, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

    get_extractor().reload_config()

    return jsonify({"success": True, "deleted": template_name})


# ================================================================
# API: 版式管理 — 应用 Few-shot 生成的配置
# ================================================================

def _sanitize_anchor_layout_detection(value):
    """Keep only the supported, serializable sample-signature shape."""
    if not isinstance(value, dict) or value.get("mode") != "anchor_layout":
        return None
    features = []
    for feature in value.get("features") or []:
        if not isinstance(feature, dict):
            continue
        text = str(feature.get("text") or "").strip()
        try:
            x = float(feature.get("x"))
            y = float(feature.get("y"))
            weight = float(feature.get("weight", 1.0))
        except (TypeError, ValueError):
            continue
        if not text or not (0.0 <= x <= 1.0 and 0.0 <= y <= 1.0):
            continue
        features.append({
            "text": text[:120],
            "x": round(x, 4),
            "y": round(y, 4),
            "weight": max(0.1, min(weight, 3.0)),
            "role": "stable_text" if feature.get("role") == "stable_text" else "field_anchor",
        })
    if not features:
        return None
    try:
        min_score = float(value.get("min_score", 0.55))
        min_matches = int(value.get("min_matches", 2))
    except (TypeError, ValueError):
        min_score = 0.55
        min_matches = 2
    return {
        "mode": "anchor_layout",
        "min_score": max(0.35, min(min_score, 0.95)),
        "min_matches": max(2, min(min_matches, len(features))),
        "features": features,
    }

@app.route("/api/config/apply", methods=["POST"])
def api_config_apply():
    """
    将 Few-shot 生成的版式配置写入 config.yaml（新自包含格式）

    Body: { "template_name": "...", "keywords": [...], "fields": {...} }
    """
    import yaml

    data = request.get_json(silent=True)
    if not data or "template_name" not in data:
        return jsonify({"error": "缺少 template_name"}), 400

    config_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "config.yaml"
    )
    if not os.path.exists(config_path):
        return jsonify({"error": "config.yaml 不存在"}), 404

    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    template_name = data["template_name"]
    keywords = data.get("keywords", [])
    fields = data.get("fields", {})
    validators = data.get("validators", {})
    source = str(data.get("source") or "fewshot").strip() or "fewshot"
    detection = _sanitize_anchor_layout_detection(data.get("detection"))
    if detection:
        keywords = []
    table_headers = [str(h).strip() for h in (data.get("table_headers") or []) if str(h).strip()]
    has_table = bool(data.get("has_table")) or bool(table_headers)

    # 构建自包含模板条目
    template_fields = {}
    for fname, fcfg in fields.items():
        entry = {
            "label": fcfg.get("label", fname),
            "anchors": fcfg.get("anchors", [fcfg.get("anchor", fname)]),
            "position": fcfg.get("position", "right"),
        }
        if fcfg.get("canonical_key"):
            entry["canonical_key"] = fcfg["canonical_key"]
        if fcfg.get("validator"):
            entry["validator"] = fcfg["validator"]
        if fcfg.get("value_pattern"):
            entry["value_pattern"] = fcfg["value_pattern"]
        if fcfg.get("multi_line"):
            entry["multi_line"] = True
        if fcfg.get("allow_shared"):
            entry["allow_shared"] = True
        if fcfg.get("search_in"):
            entry["search_in"] = fcfg["search_in"]
        template_fields[fname] = entry

    if validators:
        cfg.setdefault("validators", {})
        for vname, vcfg in validators.items():
            if isinstance(vcfg, dict):
                cfg["validators"][vname] = vcfg

    if "templates" not in cfg:
        cfg["templates"] = {}
    cfg["templates"][template_name] = {
        "keywords": keywords,
        "source": source,
        "has_table": has_table,
        "fields": template_fields,
        "output": list(fields.keys()),
    }
    if detection:
        cfg["templates"][template_name]["detection"] = detection
    if table_headers:
        cfg["templates"][template_name]["table_headers"] = table_headers

    with open(config_path, "w", encoding="utf-8") as f:
        yaml.dump(cfg, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

    get_extractor().reload_config()

    return jsonify({"success": True, "template": template_name, "fields_count": len(fields)})


# ================================================================
# API: 历史记录
# ================================================================

# ================================================================
# API: 从识别结果反哺指定版式
# ================================================================

@app.route("/api/fewshot/from-result", methods=["POST"])
def api_fewshot_from_result():
    """
    将当前结果页的最终字段/表格结构合并到指定版式。

    Body:
      {
        "job_id": "...",
        "template_name": "目标版式",
        "field_names": ["字段A", "字段B"],
        "include_table": true,
        "ai_enhance": false,
        "mode": "merge | create"
      }
    """
    import yaml

    data = request.get_json(silent=True) or {}
    job_id = str(data.get("job_id") or "").strip()
    template_name = str(data.get("template_name") or "").strip()
    field_names = data.get("field_names") or []
    include_table = bool(data.get("include_table", False))
    ai_enhance = bool(data.get("ai_enhance", False))
    mode = str(data.get("mode") or "merge").strip().lower()

    if not is_valid_job_id(job_id):
        return jsonify({"error": "invalid job_id"}), 400
    if not template_name:
        return jsonify({"error": "缺少 template_name"}), 400
    if mode not in ("merge", "create"):
        return jsonify({"error": "当前仅支持 merge 或 create 模式"}), 400
    if not field_names and not include_table:
        return jsonify({"error": "至少选择一个字段或勾选表格"}), 400

    job = load_job(job_id)
    if job is None:
        return jsonify({"error": "任务不存在"}), 404
    if job.get("result") is None:
        return jsonify({"error": "尚未识别"}), 400

    config_path = config_yaml_path()
    if not os.path.exists(config_path):
        return jsonify({"error": "config.yaml 不存在"}), 404

    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f) or {}

    templates = cfg.setdefault("templates", {})
    created = False
    if mode == "create":
        if template_name in templates:
            return jsonify({"error": f"template '{template_name}' already exists"}), 409
        templates[template_name] = {
            "keywords": [],
            "source": "feedback",
            "has_table": False,
            "enabled": True,
            "hidden": False,
            "fields": {},
            "output": [],
        }
        created = True
    if template_name not in templates:
        return jsonify({"error": f"版式 '{template_name}' 不存在"}), 404

    final_result = apply_corrections(job.get("result", {}), job.get("corrections", {}))
    final_fields = final_result.get("fields", {}) or {}
    target_template = templates[template_name]
    target_fields = target_template.setdefault("fields", {})
    output = target_template.setdefault("output", list(target_fields.keys()))

    selected_fields = [str(name).strip() for name in field_names if str(name).strip()]
    merged = []
    added = []
    updated = []
    warnings = []

    for field_name in selected_fields:
        info = final_fields.get(field_name)
        if not isinstance(info, dict):
            warnings.append(f"字段 '{field_name}' 不在当前结果中，已跳过")
            continue

        value = str(info.get("corrected") or info.get("cleaned") or info.get("value") or "").strip()
        status = info.get("status")
        if status == "not_found" and not value:
            warnings.append(f"字段 '{field_name}' 当前为空，已跳过")
            continue

        schema_name = _feedback_schema_field_name(field_name, info)
        anchors = _field_feedback_anchors(field_name, info)
        if not _field_has_reliable_anchor(field_name, info):
            warnings.append(f"字段 '{field_name}' 缺少可靠 OCR 锚点，已保存字段结构但后续可能需要手动调锚点")

        if schema_name != field_name and field_name in target_fields and schema_name not in target_fields:
            target_fields[schema_name] = target_fields.pop(field_name)
        elif schema_name != field_name and field_name in target_fields and schema_name in target_fields:
            old_entry = target_fields.pop(field_name)
            if isinstance(old_entry, dict) and isinstance(target_fields.get(schema_name), dict):
                target_fields[schema_name]["anchors"] = _merge_unique_strings(
                    target_fields[schema_name].get("anchors", []),
                    old_entry.get("anchors", []),
                    limit=8,
                )

        if schema_name in target_fields and isinstance(target_fields[schema_name], dict):
            entry = target_fields[schema_name]
            entry["label"] = schema_name if "__" in schema_name else (info.get("label") or schema_name)
            entry["anchors"] = _merge_unique_strings(entry.get("anchors", []), anchors, limit=8)
            entry.setdefault("position", info.get("position") or "right")
            if info.get("canonical_key") and not entry.get("canonical_key"):
                entry["canonical_key"] = info.get("canonical_key")
            updated.append(schema_name)
        else:
            entry = {
                "label": schema_name if "__" in schema_name else (info.get("label") or schema_name),
                "anchors": anchors,
                "position": info.get("position") or "right",
            }
            if info.get("canonical_key"):
                entry["canonical_key"] = info.get("canonical_key")
            target_fields[schema_name] = entry
            added.append(schema_name)

        if schema_name != field_name and field_name in output:
            output[:] = [schema_name if item == field_name else item for item in output]
        if schema_name not in output:
            output.append(schema_name)
        merged.append(schema_name)

    table_updated = False
    if include_table:
        table = final_result.get("table", {}) or {}
        headers = [str(h).strip() for h in table.get("headers", []) if str(h).strip()]
        if headers:
            target_template["has_table"] = True
            target_template["table_headers"] = _merge_unique_strings(
                target_template.get("table_headers", []),
                headers,
            )
            table_updated = True
        else:
            warnings.append("当前结果没有可保存的表头，未更新表格结构")

    ocr_learning = apply_ocr_feedback_learning(
        target_template,
        final_fields,
        selected_fields=selected_fields,
        blocks=load_blocks(job_id),
        warnings=warnings,
    )

    if created:
        job_blocks = load_blocks(job_id)
        detection = _feedback_layout_signature(
            job_blocks,
            target_template,
            selected_fields,
            final_fields,
            final_result.get("meta", {}).get("image_size"),
        )
        if detection.get("features"):
            target_template["keywords"] = []
            target_template["detection"] = detection
        else:
            warnings.append("当前样本没有形成可靠版式特征，新版式已保存但暂不会自动命中")

    ai_changes = {"applied": False, "keywords": [], "fields": [], "table_headers": []}
    if ai_enhance:
        ai_changes = ai_enhance_feedback_template(
            job_id=job_id,
            template_name=template_name,
            target_template=target_template,
            final_result=final_result,
            selected_fields=selected_fields,
            include_table=include_table,
            warnings=warnings,
        )

    if not merged and not table_updated:
        return jsonify({"error": "没有可反哺的字段或表格", "warnings": warnings}), 400

    target_template["output"] = output

    with open(config_path, "w", encoding="utf-8") as f:
        yaml.dump(cfg, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

    get_extractor().reload_config()

    return jsonify({
        "success": True,
        "mode": mode,
        "created": created,
        "template": template_name,
        "fields_merged": merged,
        "fields_added": added,
        "fields_updated": updated,
        "include_table": include_table,
        "table_updated": table_updated,
        "ocr_learning": ocr_learning,
        "ai_enhanced": bool(ai_changes.get("applied")),
        "ai_changes": ai_changes,
        "warnings": warnings,
    })


@app.route("/api/history", methods=["GET"])
def api_history():
    """返回所有完成过的识别任务列表（从磁盘读取，重启不丢失）"""
    history = []

    # 从 uploads 目录扫描已保存的 result.json
    uploads = app.config["UPLOAD_FOLDER"]
    if os.path.exists(uploads):
        for job_name in os.listdir(uploads):
            if not is_valid_job_id(job_name):
                continue
            job_dir_path = os.path.join(uploads, job_name)
            if not os.path.isdir(job_dir_path):
                continue
            result_path = os.path.join(job_dir_path, "result.json")
            if not os.path.exists(result_path):
                continue
            try:
                with open(result_path, "r", encoding="utf-8") as f:
                    r = json.load(f)
                m = r.get("meta", {})
                history.append({
                    "job_id": job_name,
                    "filename": r.get("filename", ""),
                    "template": r.get("template", "?"),
                    "fields_extracted": m.get("fields_extracted", 0),
                    "fields_total": m.get("fields_total", 0),
                    "ocr_blocks": m.get("ocr_blocks", 0),
                    "recognized_at": r.get("recognized_at", ""),
                })
            except Exception:
                pass

    # 也加上内存中已识别但还未写盘的任务
    for jid, job in _jobs.items():
        if job.get("status") == "done" and not any(h["job_id"] == jid for h in history):
            r = job.get("result", {})
            m = r.get("meta", {})
            history.append({
                "job_id": jid,
                "filename": job.get("filename", ""),
                "template": r.get("template", "?"),
                "fields_extracted": m.get("fields_extracted", 0),
                "fields_total": m.get("fields_total", 0),
                "ocr_blocks": m.get("ocr_blocks", 0),
                "recognized_at": r.get("recognized_at", ""),
            })

    # 按时间降序
    history.sort(key=lambda x: x.get("recognized_at", ""), reverse=True)
    return jsonify({"history": history, "count": len(history)})


@app.route("/api/history/<job_id>", methods=["DELETE"])
def api_history_delete(job_id):
    """删除单条历史记录（磁盘 + 内存）"""
    deleted = False

    # 从内存移除
    if job_id in _jobs:
        del _jobs[job_id]
        deleted = True

    # 从磁盘删除
    target_dir = job_dir(job_id, create=False)
    if os.path.exists(target_dir):
        shutil.rmtree(target_dir, ignore_errors=True)
        deleted = True

    if deleted:
        return jsonify({"success": True, "deleted": job_id})
    return jsonify({"error": "记录不存在"}), 404


@app.route("/api/history", methods=["DELETE"])
def api_history_delete_all():
    """删除全部历史记录（磁盘 + 内存）"""
    count = 0

    # 清空所有已完成任务（保留可能正在处理中的？全删即可）
    done_ids = [jid for jid, j in _jobs.items() if j.get("status") == "done"]
    for jid in done_ids:
        del _jobs[jid]
        count += 1

    # 删除 uploads 目录下所有子目录
    uploads = app.config["UPLOAD_FOLDER"]
    if os.path.exists(uploads):
        for name in os.listdir(uploads):
            if not is_valid_job_id(name):
                continue
            dir_path = os.path.join(uploads, name)
            if os.path.isdir(dir_path) and os.path.exists(os.path.join(dir_path, "result.json")):
                shutil.rmtree(dir_path, ignore_errors=True)
                count += 1

    return jsonify({"success": True, "deleted_count": count})


# ================================================================
# API: Few-shot 版式学习
# ================================================================

@app.route("/api/fewshot/learn", methods=["POST"])
def api_fewshot_learn():
    """
    Few-shot 版式学习：上传 1~5 份 PDF + GT JSON → 自动生成配置

    Body (multipart/form-data):
      - files: PDF 文件列表
      - gts:   JSON 字符串数组 (每个对应一份 PDF)
    """
    from fewshot import FewShotLearner

    if "files" not in request.files:
        return jsonify({"error": "缺少 files"}), 400

    files = request.files.getlist("files")
    gt_strs = request.form.getlist("gts")
    ai_enhance = str(request.form.get("ai_enhance") or "").strip().lower() in {"1", "true", "yes", "on"}

    if len(files) != len(gt_strs):
        return jsonify({"error": "files 和 gts 数量不匹配"}), 400
    if len(files) < 1:
        return jsonify({"error": "至少需要 1 份样本"}), 400
    if len(files) > 5:
        return jsonify({"error": "最多 5 份样本"}), 400

    # 保存临时文件
    tmp_dir = tempfile.mkdtemp()
    samples = []
    try:
        for i, (f, gt_str) in enumerate(zip(files, gt_strs)):
            ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else "pdf"
            tmp_path = os.path.join(tmp_dir, f"sample_{i}.{ext}")
            f.save(tmp_path)
            gt = json.loads(gt_str)
            samples.append((tmp_path, gt))

        learner = FewShotLearner()
        result = learner.learn(samples)
        result.setdefault("ai_enhanced", False)
        result.setdefault("ai_changes", {"applied": False, "keywords": [], "fields": [], "table_headers": []})
        warnings = []
        if ai_enhance:
            ai_changes = ai_enhance_fewshot_learning(samples, result, warnings)
            result["ai_enhanced"] = bool(ai_changes.get("applied"))
            result["ai_changes"] = ai_changes
        if warnings:
            result["warnings"] = warnings

        return jsonify({
            "success": True,
            "result": result,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ================================================================
# API: 视觉兜底模型设置
# ================================================================

@app.route("/api/vision-settings", methods=["GET"])
def api_vision_settings_get():
    """获取视觉兜底设置；API Key 只返回掩码，不回传明文。"""
    return jsonify({
        "settings": load_vision_settings(include_secret=False),
        "options": vision_settings_options(),
    })


@app.route("/api/vision-settings", methods=["POST"])
def api_vision_settings_save():
    """保存视觉兜底设置，包含供应商、模型、阈值和可选 API Key。"""
    data = request.get_json(silent=True) or {}
    settings = save_vision_settings(data)
    return jsonify({
        "success": True,
        "settings": settings,
        "options": vision_settings_options(),
    })


@app.route("/api/vision-settings/probe", methods=["POST"])
def api_vision_settings_probe():
    """按当前表单配置检测可用模型；不保存 API Key 或模型设置。"""
    data = request.get_json(silent=True) or {}
    result = probe_vision_models(data, saved_settings=load_vision_settings(include_secret=True))
    return jsonify(result), 200 if result.get("success") else 400


@app.route("/api/vision-settings/api-key", methods=["GET"])
def api_vision_settings_api_key():
    """按需返回本地保存的 API Key 明文；仅用于本地设置弹窗的小眼睛查看。"""
    settings = load_vision_settings(include_secret=True)
    provider = _normalize_vision_provider(request.args.get("provider") or settings.get("provider"), request.args.get("base_url") or settings.get("base_url"))
    model = normalize_vision_model(provider, request.args.get("model") or settings.get("model"))
    profile = _normalize_vision_profile(provider, (settings.get("profiles") or {}).get(provider))
    model_api_keys = profile.get("model_api_keys") or {}
    api_key = str(model_api_keys.get(model) or profile.get("api_key") or "")
    return jsonify({
        "success": True,
        "has_api_key": bool(api_key),
        "api_key": api_key,
    })


@app.route("/api/vision-settings", methods=["DELETE"])
def api_vision_settings_clear():
    """清除本地保存的视觉兜底设置和 API Key，恢复默认千问配置。"""
    settings = clear_vision_settings()
    return jsonify({
        "success": True,
        "settings": settings,
        "options": vision_settings_options(),
    })


# ================================================================
# API: 健康检查
# ================================================================

@app.route("/api/health", methods=["GET"])
def api_health():
    return jsonify({
        "status": "ok",
        "active_jobs": len(_jobs),
    })


@app.route("/api/jobs", methods=["GET"])
def api_jobs():
    """列出所有任务状态"""
    summary = []
    for jid, job in _jobs.items():
        summary.append({
            "job_id": jid,
            "filename": job.get("filename", ""),
            "status": job.get("status", "unknown"),
            "created_at": job.get("created_at", ""),
        })
    return jsonify({"jobs": summary})


# ================================================================
# 启动
# ================================================================

if __name__ == "__main__":
    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
    print("\n  SmartLDS API Server")
    print(f"  上传目录: {app.config['UPLOAD_FOLDER']}")
    print(f"  最大文件: {app.config['MAX_CONTENT_LENGTH'] // (1024*1024)}MB")
    print("\n  端点:")
    print("    POST /api/upload          上传文件")
    print("    POST /api/recognize/<id>  触发识别")
    print("    GET  /api/result/<id>     获取结果")
    print("    POST /api/correct/<id>    人工校正")
    print("    GET  /api/export/<id>     导出 (format=json|xlsx)")
    print("    GET  /api/image/<id>      获取图片")
    print("    GET  /api/health          健康检查")
    print()
    app.run(host="0.0.0.0", port=5000, debug=True)
